from PyQt5.QtGui import QTextDocument, QTextCursor
from PyQt5.QtPrintSupport import QPrintPreviewDialog
from openpyxl import Workbook, load_workbook

# wb = Workbook()
# ws = wb.active
# ws1 = wb.create_sheet('mysheet',0)
# print(wb.sheetnames)

# wb.save('balances.xlsx')

def handlePaintRequest(self, printer):
    document = QTextDocument()
    cursor = QTextCursor(document)
    cursor.insertText(self.label.text())
    document.print(printer)

wb = load_workbook('blances.xlsx')
print(wb.sheetnames)
ws=wb["mysheet"]
print(ws['A4'].value)
dialog = QPrintPreviewDialog()
dialog.paintRequested.connect(handlePaintRequest)
dialog.exec_()



